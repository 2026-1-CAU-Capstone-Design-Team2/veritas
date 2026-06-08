from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from core.local_corpus_models import LocalFileManifestEntry

# Tabular extensions this service can query directly.
TABLE_EXTENSIONS = {".csv", ".xlsx"}
# Cap on rows returned to the LLM so a broad query cannot flood its context.
MAX_RESULT_ROWS = 200
DEFAULT_RESULT_ROWS = 50
# Sample rows included by describe().
DESCRIBE_SAMPLE_ROWS = 5

FILTER_OPS = ("==", "!=", ">", "<", ">=", "<=", "contains")
AGGREGATE_FUNCS = ("sum", "mean", "min", "max", "count")

# Numeric part of values like "45,000원", "약 1,234.5 USD", "-3%".
_NUMBER_RE = re.compile(r"-?\d+(?:,\d{3})*(?:\.\d+)?")


class TableQueryError(ValueError):
    """User-facing error for invalid table queries (unknown table, column, op...)."""


def _cell_to_text(cell: Any) -> str:
    if cell is None:
        return ""
    if isinstance(cell, bool):
        return str(cell)
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    if hasattr(cell, "isoformat"):
        return cell.isoformat()
    return str(cell).strip()


def _to_number(value: Any) -> float | None:
    """Best-effort numeric coercion for table cells and filter values.

    Handles thousand separators and unit suffixes ("45,000원", "-3%"). Values
    whose numeric part covers less than half the text (e.g. dates like
    "2025-03-15") are NOT treated as numbers so they fall back to string
    comparison, where ISO dates order correctly.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        pass
    match = _NUMBER_RE.search(text)
    if not match:
        return None
    content_len = len(re.sub(r"\s", "", text))
    if len(match.group(0)) * 2 < content_len:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


class TableQueryService:
    """Loss-free structured queries over registered local .csv/.xlsx files.

    Reads the ORIGINAL files listed in the local-corpus manifest at query
    time — every row, no embeddings, no row caps from the profiling pass.
    The indexed table profile (RAG) remains the discovery surface; actual
    values always come from this service.
    """

    def __init__(self, workspace_root: str | Path) -> None:
        self._workspace_root = Path(workspace_root)

    # -- discovery -----------------------------------------------------------

    def list_tables(self) -> dict[str, Any]:
        tables: list[dict[str, Any]] = []
        for entry in self._table_entries():
            info: dict[str, Any] = {
                "file_name": entry.file_name,
                "source_id": entry.source_id,
                "extension": entry.extension,
                "relative_path": entry.relative_path,
            }
            try:
                if entry.extension.lower() == ".xlsx":
                    info["sheets"] = [
                        {"sheet_name": name, "columns": columns}
                        for name, columns in self._xlsx_sheet_headers(entry.absolute_path)
                    ]
                else:
                    info["columns"] = self._csv_header(entry.absolute_path)
            except Exception as exc:
                info["error"] = str(exc)[:200]
            tables.append(info)
        return {"tables": tables, "table_count": len(tables)}

    def describe(self, file_name: str, sheet_name: str | None = None) -> dict[str, Any]:
        entry = self._resolve_entry(file_name)
        header, rows, resolved_sheet = self._load_rows(entry, sheet_name)
        columns: list[dict[str, Any]] = []
        for index, name in enumerate(header):
            values = [row[index] for row in rows if index < len(row)]
            numbers = [n for n in (_to_number(v) for v in values if v != "") if n is not None]
            non_empty = [v for v in values if v != ""]
            inferred = "number" if non_empty and len(numbers) >= max(1, len(non_empty) // 2) else "text"
            column_info: dict[str, Any] = {"name": name, "inferred_type": inferred}
            if numbers:
                column_info["min"] = min(numbers)
                column_info["max"] = max(numbers)
            columns.append(column_info)
        return {
            "file_name": entry.file_name,
            "sheet_name": resolved_sheet,
            "total_rows": len(rows),
            "columns": columns,
            "sample_rows": [
                dict(zip(header, row)) for row in rows[:DESCRIBE_SAMPLE_ROWS]
            ],
        }

    # -- query ----------------------------------------------------------------

    def query(
        self,
        file_name: str,
        *,
        sheet_name: str | None = None,
        columns: list[str] | None = None,
        where: list[dict[str, Any]] | None = None,
        group_by: list[str] | None = None,
        aggregate: list[dict[str, Any]] | None = None,
        sort_by: str | None = None,
        descending: bool = False,
        limit: int = DEFAULT_RESULT_ROWS,
    ) -> dict[str, Any]:
        entry = self._resolve_entry(file_name)
        header, rows, resolved_sheet = self._load_rows(entry, sheet_name)
        total_rows = len(rows)

        row_maps = [dict(zip(header, row)) for row in rows]
        conditions = self._validate_conditions(where or [], header)
        matched = [row for row in row_maps if self._row_matches(row, conditions)]

        if aggregate:
            result_rows = self._aggregate_rows(matched, group_by or [], aggregate, header)
        else:
            selected = self._resolve_columns(columns or [], header) if columns else header
            result_rows = [{name: row.get(name, "") for name in selected} for row in matched]

        if sort_by:
            sort_column = self._resolve_column_in_rows(sort_by, result_rows, header)
            result_rows.sort(
                key=lambda row: self._sort_key(row.get(sort_column, "")),
                reverse=bool(descending),
            )

        limit = max(1, min(MAX_RESULT_ROWS, int(limit or DEFAULT_RESULT_ROWS)))
        truncated = len(result_rows) > limit
        return {
            "file_name": entry.file_name,
            "sheet_name": resolved_sheet,
            "total_rows": total_rows,
            "matched_rows": len(matched),
            "returned_rows": min(len(result_rows), limit),
            "truncated": truncated,
            "rows": result_rows[:limit],
        }

    # -- manifest / file loading ----------------------------------------------

    def _table_entries(self) -> list[LocalFileManifestEntry]:
        manifest_path = self._workspace_root / "local" / "manifest.json"
        if not manifest_path.exists():
            return []
        try:
            payload = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            return []
        entries: list[LocalFileManifestEntry] = []
        for item in payload if isinstance(payload, list) else []:
            if not isinstance(item, dict):
                continue
            try:
                entry = LocalFileManifestEntry(**item)
            except TypeError:
                continue
            # Querying does not depend on parse/index success — even files the
            # indexer skipped (e.g. too large) remain queryable from the original.
            if entry.extension.lower() not in TABLE_EXTENSIONS:
                continue
            if not Path(entry.absolute_path).exists():
                continue
            entries.append(entry)
        return entries

    def _resolve_entry(self, file_name: str) -> LocalFileManifestEntry:
        requested = str(file_name or "").strip()
        if not requested:
            raise TableQueryError("`file_name` is required. Call list_tables first.")
        entries = self._table_entries()
        if not entries:
            raise TableQueryError(
                "No local .csv/.xlsx files are registered. "
                "Register a local access folder containing table files first."
            )
        lowered = requested.lower()
        for entry in entries:
            if entry.file_name == requested or entry.source_id == requested:
                return entry
        for entry in entries:
            if entry.file_name.lower() == lowered:
                return entry
        for entry in entries:
            if entry.relative_path.lower().endswith(lowered):
                return entry
        available = ", ".join(sorted({entry.file_name for entry in entries}))
        raise TableQueryError(f"Unknown table file: {requested}. Available: {available}")

    def _load_rows(
        self,
        entry: LocalFileManifestEntry,
        sheet_name: str | None,
    ) -> tuple[list[str], list[list[str]], str | None]:
        """Read the FULL table: header names, all data rows, resolved sheet name."""
        if entry.extension.lower() == ".xlsx":
            return self._load_xlsx_rows(entry.absolute_path, sheet_name)
        header, rows = self._load_csv_rows(entry.absolute_path)
        return header, rows, None

    def _load_csv_rows(self, path: str) -> tuple[list[str], list[list[str]]]:
        raw_rows: list[list[str]] = []
        for encoding in ("utf-8-sig", "cp949"):
            try:
                with Path(path).open("r", encoding=encoding, newline="") as handle:
                    raw_rows = [list(row) for row in csv.reader(handle)]
                break
            except UnicodeDecodeError:
                continue
        else:
            with Path(path).open("r", encoding="utf-8", errors="replace", newline="") as handle:
                raw_rows = [list(row) for row in csv.reader(handle)]
        return self._normalize_rows(raw_rows)

    def _load_xlsx_rows(
        self,
        path: str,
        sheet_name: str | None,
    ) -> tuple[list[str], list[list[str]], str]:
        try:
            import openpyxl  # type: ignore
        except Exception as exc:
            raise TableQueryError("openpyxl is required to query .xlsx files") from exc

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = self._resolve_sheet(workbook, sheet_name)
            raw_rows = [
                [_cell_to_text(cell) for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            header, rows = self._normalize_rows(raw_rows)
            return header, rows, sheet.title
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def _resolve_sheet(self, workbook, sheet_name: str | None):
        names = workbook.sheetnames
        if not names:
            raise TableQueryError("The workbook has no sheets.")
        requested = str(sheet_name or "").strip()
        if not requested:
            return workbook[names[0]]
        for name in names:
            if name == requested or name.lower() == requested.lower():
                return workbook[name]
        raise TableQueryError(f"Unknown sheet: {requested}. Available: {', '.join(names)}")

    def _normalize_rows(self, raw_rows: list[list[Any]]) -> tuple[list[str], list[list[str]]]:
        rows = [[_cell_to_text(cell) for cell in row] for row in raw_rows]
        rows = [row for row in rows if any(cell != "" for cell in row)]
        if not rows:
            return [], []
        header = [
            str(cell or "").strip() or f"column_{index + 1}"
            for index, cell in enumerate(rows[0])
        ]
        width = len(header)
        data_rows = [
            (row + [""] * max(0, width - len(row)))[:width]
            for row in rows[1:]
        ]
        return header, data_rows

    def _csv_header(self, path: str) -> list[str]:
        # Header-only read so list_tables / the chat table catalog stay cheap on
        # large CSVs (the full-file _load_csv_rows is reserved for describe/query).
        first_row = self._read_first_csv_row(path)
        return [
            str(cell or "").strip() or f"column_{index + 1}"
            for index, cell in enumerate(first_row)
        ]

    def _read_first_csv_row(self, path: str) -> list[str]:
        def _first_nonempty(handle) -> list[str]:
            for row in csv.reader(handle):
                if any(_cell_to_text(cell) for cell in row):
                    return list(row)
            return []

        for encoding in ("utf-8-sig", "cp949"):
            try:
                with Path(path).open("r", encoding=encoding, newline="") as handle:
                    return _first_nonempty(handle)
            except UnicodeDecodeError:
                continue
        with Path(path).open("r", encoding="utf-8", errors="replace", newline="") as handle:
            return _first_nonempty(handle)

    def _xlsx_sheet_headers(self, path: str) -> list[tuple[str, list[str]]]:
        try:
            import openpyxl  # type: ignore
        except Exception as exc:
            raise TableQueryError("openpyxl is required to query .xlsx files") from exc

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            headers: list[tuple[str, list[str]]] = []
            for sheet in workbook.worksheets:
                first_row: list[str] = []
                for row in sheet.iter_rows(values_only=True, max_row=1):
                    first_row = [
                        str(_cell_to_text(cell)).strip() or f"column_{index + 1}"
                        for index, cell in enumerate(row)
                    ]
                    break
                headers.append((sheet.title, first_row))
            return headers
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    # -- filtering / aggregation ----------------------------------------------

    def _validate_conditions(
        self,
        where: list[dict[str, Any]],
        header: list[str],
    ) -> list[dict[str, Any]]:
        conditions: list[dict[str, Any]] = []
        for raw in where:
            if not isinstance(raw, dict):
                continue
            column = self._resolve_column(str(raw.get("column") or ""), header)
            op = str(raw.get("op") or "==").strip()
            if op not in FILTER_OPS:
                raise TableQueryError(
                    f"Unsupported operator: {op}. Supported: {', '.join(FILTER_OPS)}"
                )
            conditions.append({"column": column, "op": op, "value": raw.get("value", "")})
        return conditions

    def _resolve_column(self, name: str, header: list[str]) -> str:
        requested = str(name or "").strip()
        if requested in header:
            return requested
        lowered = requested.lower()
        for column in header:
            if column.lower() == lowered:
                return column
        raise TableQueryError(
            f"Unknown column: {requested}. Available: {', '.join(header)}"
        )

    def _resolve_columns(self, names: list[str], header: list[str]) -> list[str]:
        return [self._resolve_column(name, header) for name in names]

    def _resolve_column_in_rows(
        self,
        name: str,
        rows: list[dict[str, Any]],
        header: list[str],
    ) -> str:
        """Resolve a sort column against result rows (which may hold aggregate
        labels like "sum(매출)") falling back to the source header."""
        requested = str(name or "").strip()
        if rows and requested in rows[0]:
            return requested
        if rows:
            lowered = requested.lower()
            for key in rows[0]:
                if key.lower() == lowered:
                    return key
        return self._resolve_column(requested, header)

    def _row_matches(self, row: dict[str, str], conditions: list[dict[str, Any]]) -> bool:
        for condition in conditions:
            actual = row.get(condition["column"], "")
            if not self._compare(actual, condition["op"], condition["value"]):
                return False
        return True

    def _compare(self, actual: Any, op: str, expected: Any) -> bool:
        actual_text = _cell_to_text(actual)
        expected_text = _cell_to_text(expected)
        if op == "contains":
            return expected_text.lower() in actual_text.lower()

        actual_number = _to_number(actual)
        expected_number = _to_number(expected)
        if actual_number is not None and expected_number is not None:
            left: Any = actual_number
            right: Any = expected_number
        else:
            left = actual_text.lower()
            right = expected_text.lower()

        if op == "==":
            return left == right
        if op == "!=":
            return left != right
        if op == ">":
            return left > right
        if op == "<":
            return left < right
        if op == ">=":
            return left >= right
        if op == "<=":
            return left <= right
        raise TableQueryError(f"Unsupported operator: {op}")

    def _aggregate_rows(
        self,
        rows: list[dict[str, str]],
        group_by: list[str],
        aggregates: list[dict[str, Any]],
        header: list[str],
    ) -> list[dict[str, Any]]:
        group_columns = self._resolve_columns(group_by, header) if group_by else []
        specs: list[tuple[str, str]] = []
        for raw in aggregates:
            if not isinstance(raw, dict):
                continue
            func = str(raw.get("func") or "").strip().lower()
            if func not in AGGREGATE_FUNCS:
                raise TableQueryError(
                    f"Unsupported aggregate: {func}. Supported: {', '.join(AGGREGATE_FUNCS)}"
                )
            column = str(raw.get("column") or "*").strip()
            if func != "count" or column not in ("", "*"):
                column = self._resolve_column(column, header)
            specs.append((column, func))
        if not specs:
            raise TableQueryError("At least one aggregate {column, func} is required.")

        groups: dict[tuple, list[dict[str, str]]] = {}
        for row in rows:
            key = tuple(row.get(column, "") for column in group_columns)
            groups.setdefault(key, []).append(row)
        if not groups and not group_columns:
            groups[()] = []

        result: list[dict[str, Any]] = []
        for key, group_rows in groups.items():
            out: dict[str, Any] = dict(zip(group_columns, key))
            for column, func in specs:
                label = f"{func}({column or '*'})"
                out[label] = self._compute_aggregate(group_rows, column, func)
            result.append(out)
        return result

    def _compute_aggregate(
        self,
        rows: list[dict[str, str]],
        column: str,
        func: str,
    ) -> float | int | None:
        if func == "count":
            if column in ("", "*"):
                return len(rows)
            return sum(1 for row in rows if _cell_to_text(row.get(column, "")) != "")
        numbers = [
            number
            for number in (_to_number(row.get(column, "")) for row in rows)
            if number is not None
        ]
        if not numbers:
            return None
        if func == "sum":
            return round(sum(numbers), 6)
        if func == "mean":
            return round(sum(numbers) / len(numbers), 6)
        if func == "min":
            return min(numbers)
        if func == "max":
            return max(numbers)
        raise TableQueryError(f"Unsupported aggregate: {func}")

    def _sort_key(self, value: Any) -> tuple[int, Any]:
        number = _to_number(value)
        if number is not None:
            return (0, number)
        return (1, _cell_to_text(value).lower())


__all__ = [
    "AGGREGATE_FUNCS",
    "DEFAULT_RESULT_ROWS",
    "FILTER_OPS",
    "MAX_RESULT_ROWS",
    "TABLE_EXTENSIONS",
    "TableQueryError",
    "TableQueryService",
]
